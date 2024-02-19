import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor
from json import JSONDecodeError

import openpyxl
from dotenv import load_dotenv

import log_util
from gpt_util import OpenAIClient

logger = log_util.get_logger(__name__)

load_dotenv()
# GPT3.5
AZURE_OPENAI_MODEL_GPT3_5 = os.getenv("AZURE_OPENAI_MODEL_GPT3_5")
AZURE_OPENAI_MODEL_GPT3_5_API_BASE = os.getenv("AZURE_OPENAI_MODEL_GPT3_5_API_BASE")
AZURE_OPENAI_MODEL_GPT3_5_API_VERSION = os.getenv("AZURE_OPENAI_MODEL_GPT3_5_API_VERSION")
AZURE_OPENAI_MODEL_GPT3_5_API_KEY = os.getenv("AZURE_OPENAI_MODEL_GPT3_5_API_KEY")
# GPT4
AZURE_OPENAI_MODEL_GPT4 = os.getenv("AZURE_OPENAI_MODEL_GPT4")
AZURE_OPENAI_MODEL_GPT4_API_BASE = os.getenv("AZURE_OPENAI_MODEL_GPT4_API_BASE")
AZURE_OPENAI_MODEL_GPT4_API_VERSION = os.getenv("AZURE_OPENAI_MODEL_GPT4_API_VERSION")
AZURE_OPENAI_MODEL_GPT4_API_KEY = os.getenv("AZURE_OPENAI_MODEL_GPT4_API_KEY")
# EMBEDDING
AZURE_OPENAI_MODEL_EMBEDDING = os.getenv("AZURE_OPENAI_MODEL_EMBEDDING")
AZURE_OPENAI_MODEL_EMBEDDING_API_BASE = os.getenv("AZURE_OPENAI_MODEL_EMBEDDING_API_BASE")
AZURE_OPENAI_MODEL_EMBEDDING_API_VERSION = os.getenv("AZURE_OPENAI_MODEL_EMBEDDING_API_VERSION")
AZURE_OPENAI_MODEL_EMBEDDING_API_KEY = os.getenv("AZURE_OPENAI_MODEL_EMBEDDING_API_KEY")
# VECTOR_DATABASE_PINECONE
VECTOR_DATABASE_PINECONE_INDEX = os.getenv("VECTOR_DATABASE_PINECONE_INDEX")
VECTOR_DATABASE_PINECONE_ENVIRONMENT = os.getenv("VECTOR_DATABASE_PINECONE_ENVIRONMENT")
VECTOR_DATABASE_PINECONE_API_KEY = os.getenv("VECTOR_DATABASE_PINECONE_API_KEY")

prompt = open('src/prompt/prompt', 'r')
PROMPT_CONTENT = prompt.read()
prompt.close()


def get_gpt_instance(enable_gpt4=False):
    if enable_gpt4 is not None and (enable_gpt4 is True or enable_gpt4 == 'true'):
        gpt_client = OpenAIClient(
            api_type="azure",
            api_base=AZURE_OPENAI_MODEL_GPT4_API_BASE,
            api_version=AZURE_OPENAI_MODEL_GPT4_API_VERSION,
            api_key=AZURE_OPENAI_MODEL_GPT4_API_KEY,
            chatModel=AZURE_OPENAI_MODEL_GPT4
        )
    else:
        gpt_client = OpenAIClient(
            api_type="azure",
            api_base=AZURE_OPENAI_MODEL_GPT3_5_API_BASE,
            api_version=AZURE_OPENAI_MODEL_GPT3_5_API_VERSION,
            api_key=AZURE_OPENAI_MODEL_GPT3_5_API_KEY,
            chatModel=AZURE_OPENAI_MODEL_GPT3_5
        )
    return gpt_client


# 这个函数用于将一个列表中的内容进行翻译，返回一个翻译后的列表
# 其中，列表的第一个元素是分类，第二个元素是标题，第三个元素是url(没有用到)，第四个元素是正文
# 返回的列表中，前四个元素是原来的内容，后面的元素是翻译后的内容，分别是波兰语(pl)，德语(de)，俄语(ru)，法语(fr)，韩语(ko)，荷兰语(nl)，日语(ja)，土耳其语(tr)，西班牙语(es)，意大利语(it)，越南语(vi)
# 如果翻译失败，那么就会返回None
# 由于GPT-3的限制，每次翻译的内容不能超过300个token，会调用长句翻译，将正文分成多个部分，然后分别翻译，最后再将翻译后的内容合并起来，因此可以处理大文本
# 但是GPT-3在进行大文本翻译的时候非常非常慢
# 但是由于gpt类中split_text函数的限制，如果一个单句的token数量超过1800（预设值，可修改），则会出现错误返回None
# input args:
#  row_number: 行号,用来表示在input excel中的行位置，这里仅用户打印
#  row_data: 一行输入数据，是一个列表，包含了分类，标题，url，正文
#  min_interval: 最小间隔，如果翻译时间小于min_interval，那么会sleep，这是由于GPT调用时往往会在一定时间内有次数限制，如果短时间内调用次数过多，会返回错误
def process_row(row_number, row_data, target_languages, progress_callback=None, min_interval=1, enable_gpt4=False):
    title, label, _, text = row_data
    if not text:
        return row_data + (None,) * len(target_languages)

    start_time = time.time()

    gpt_client = get_gpt_instance(enable_gpt4)
    logger.info(f"processing row: {row_number}")
    if progress_callback:
        progress_callback(f"processing string: {row_number}")

    try:
        text = text.replace('"', "'").replace("\n", "")
        target_languages_str = ', '.join(target_languages)
        trans_prompt = PROMPT_CONTENT.replace("{target_languages}", target_languages_str).replace("{origin_text}", text)

        translations_jsonstr = gpt_client.chatCompletionsCreate(
            temperature=0.0,
            messages=[{'role': 'user', 'content': trans_prompt}]
        )

        try:
            translations_json = json.loads(translations_jsonstr)
        except JSONDecodeError:
            logger.error(f"Json解析异常, translations_jsonstr: {translations_jsonstr}")
            return row_data

        translated_texts = []
        for target_language in target_languages:
            target_language = re.compile(r"\((.*?)\)").search(target_language).group(1)
            if target_language in translations_json:
                translated_texts.append(translations_json[target_language])
            else:
                translated_texts.append("")

        return row_data + tuple(translated_texts)
    except Exception as e:
        logger.error(f"捕获到异常:{type(e).__name__}", exc_info=True)
        return row_data
    finally:
        end_time = time.time()
        elapsed_time = end_time - start_time

        if elapsed_time < min_interval:
            time.sleep(min_interval - elapsed_time)

        logger.info(f"row {row_number} processed in {elapsed_time} seconds")
        if progress_callback:
            progress_callback(f"string {row_number} processed in {elapsed_time} seconds")


def process_excel(input_file, output_file, min_row=2, max_row=None, target_languages=None, progress_callback=None,
                  enable_gpt4=False):
    input_wb = openpyxl.load_workbook(input_file)
    input_sheet = input_wb.active

    # Copy headers from input sheet to output sheet
    headers = [cell for cell in next(input_sheet.iter_rows(min_row=1, max_row=1, max_col=4, values_only=True))]
    headers += target_languages
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.append(headers)

    max_row = input_sheet.max_row if max_row is None else max_row
    # Process each row in the input sheet using ThreadPoolExecutor
    input_data = list(
        input_sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=1,
            max_col=4,
            values_only=True
        )
    )

    # q: explain the code below
    # a: https://stackoverflow.com/questions/6893968/how-to-get-the-return-value-from-a-thread-in-python
    row_count = 0
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {
            executor.submit(
                process_row, row_number, row_data, target_languages, enable_gpt4=enable_gpt4
            ) for row_number, row_data in enumerate(input_data, start=min_row)
        }

        for future in futures:
            row_count = row_count + 1
            if progress_callback:
                progress_callback(f"{row_count}/{max_row} rows processed")
            output_data = future.result()
            output_sheet.append(output_data)
            output_wb.save(output_file)

    if progress_callback:
        progress_callback('All translation finished.')

    input_wb.close()
    output_wb.close()
